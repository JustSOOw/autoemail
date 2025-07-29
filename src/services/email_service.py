# -*- coding: utf-8 -*-
"""
域名邮箱管理器 - 简化邮箱服务
专注于邮箱生成、存储和管理的核心功能
"""

import json
from datetime import datetime
from typing import List, Optional, Dict, Any

from models.email_model import EmailModel, EmailStatus, create_email_model
from models.config_model import ConfigModel
from services.database_service import DatabaseService
from services.email_generator import EmailGenerator
from utils.logger import get_logger


class EmailService:
    """
    简化邮箱服务类
    
    提供邮箱生成、存储和管理的核心功能
    """

    def __init__(self, config: ConfigModel, db_service: DatabaseService):
        """
        初始化邮箱服务
        
        Args:
            config: 配置模型实例
            db_service: 数据库服务实例
        """
        self.config = config
        self.db_service = db_service
        self.logger = get_logger(__name__)
        
        # 初始化邮箱生成器
        self.email_generator = EmailGenerator(config)
        
        self.logger.info("简化邮箱服务初始化完成")

    def create_email(self, 
                    prefix_type: str = "random_name",
                    custom_prefix: Optional[str] = None,
                    tags: Optional[List[str]] = None,
                    notes: str = "") -> EmailModel:
        """
        创建新邮箱
        
        Args:
            prefix_type: 前缀类型
            custom_prefix: 自定义前缀
            tags: 标签列表
            notes: 备注
            
        Returns:
            创建的邮箱模型
        """
        try:
            # 生成邮箱地址
            email_address = self.email_generator.generate_email(
                prefix_type=prefix_type,
                custom_prefix=custom_prefix,
                add_timestamp=True
            )
            
            # 创建邮箱模型
            email_model = create_email_model(
                email_address=email_address,
                tags=tags or [],
                notes=notes
            )
            
            # 保存到数据库
            email_id = self._save_email_to_db(email_model)
            email_model.id = email_id
            
            self.logger.info(f"成功创建邮箱: {email_address}")
            return email_model
            
        except Exception as e:
            self.logger.error(f"创建邮箱失败: {e}")
            raise

    def get_email_by_id(self, email_id: int) -> Optional[EmailModel]:
        """
        根据ID获取邮箱
        
        Args:
            email_id: 邮箱ID
            
        Returns:
            邮箱模型或None
        """
        try:
            query = "SELECT * FROM emails WHERE id = ? AND is_active = 1"
            result = self.db_service.execute_query(query, (email_id,), fetch_one=True)
            
            if result:
                return self._row_to_email_model(result)
            return None
            
        except Exception as e:
            self.logger.error(f"获取邮箱失败: {e}")
            return None

    def search_emails(self, 
                     keyword: str = "",
                     domain: str = "",
                     status: Optional[EmailStatus] = None,
                     tags: Optional[List[str]] = None,
                     limit: int = 100) -> List[EmailModel]:
        """
        搜索邮箱
        
        Args:
            keyword: 关键词
            domain: 域名
            status: 状态
            tags: 标签列表
            limit: 限制数量
            
        Returns:
            邮箱模型列表
        """
        try:
            # 构建安全的查询条件
            base_query = "SELECT * FROM emails WHERE is_active = 1"
            where_conditions = []
            params = []

            if keyword:
                where_conditions.append("(email_address LIKE ? OR notes LIKE ?)")
                params.extend([f"%{keyword}%", f"%{keyword}%"])

            if domain:
                where_conditions.append("domain = ?")
                params.append(domain)

            if status:
                where_conditions.append("status = ?")
                params.append(status.value)

            # 标签搜索需要JOIN操作
            if tags:
                for tag in tags:
                    where_conditions.append("EXISTS (SELECT 1 FROM email_tags et JOIN tags t ON et.tag_id = t.id WHERE et.email_id = emails.id AND t.name = ?)")
                    params.append(tag)

            # 安全地构建完整查询
            if where_conditions:
                query = f"{base_query} AND {' AND '.join(where_conditions)} ORDER BY created_at DESC LIMIT ?"
            else:
                query = f"{base_query} ORDER BY created_at DESC LIMIT ?"

            params.append(limit)
            
            results = self.db_service.execute_query(query, tuple(params))
            return [self._row_to_email_model(row) for row in results or []]
            
        except Exception as e:
            self.logger.error(f"搜索邮箱失败: {e}")
            return []

    def update_email(self, email_model: EmailModel) -> bool:
        """
        更新邮箱信息
        
        Args:
            email_model: 邮箱模型
            
        Returns:
            是否更新成功
        """
        try:
            email_model.updated_at = datetime.now()
            return self._update_email_in_db(email_model)
            
        except Exception as e:
            self.logger.error(f"更新邮箱失败: {e}")
            return False

    def delete_email(self, email_id: int) -> bool:
        """
        删除邮箱（软删除）
        
        Args:
            email_id: 邮箱ID
            
        Returns:
            是否删除成功
        """
        try:
            query = "UPDATE emails SET is_active = 0, updated_at = ? WHERE id = ?"
            affected_rows = self.db_service.execute_update(
                query, 
                (datetime.now().isoformat(), email_id)
            )
            
            success = affected_rows > 0
            if success:
                self.logger.info(f"成功删除邮箱: {email_id}")
            
            return success
            
        except Exception as e:
            self.logger.error(f"删除邮箱失败: {e}")
            return False

    def add_email_tag(self, email_id: int, tag_name: str) -> bool:
        """
        为邮箱添加标签

        Args:
            email_id: 邮箱ID
            tag_name: 标签名称

        Returns:
            是否添加成功
        """
        try:
            with self.db_service.get_cursor() as cursor:
                self._save_email_tags(cursor, email_id, [tag_name])
            return True

        except Exception as e:
            self.logger.error(f"添加邮箱标签失败: {e}")
            return False

    def remove_email_tag(self, email_id: int, tag_name: str) -> bool:
        """
        移除邮箱标签

        Args:
            email_id: 邮箱ID
            tag_name: 标签名称

        Returns:
            是否移除成功
        """
        try:
            query = """
                DELETE FROM email_tags
                WHERE email_id = ? AND tag_id = (
                    SELECT id FROM tags WHERE name = ?
                )
            """
            affected_rows = self.db_service.execute_update(query, (email_id, tag_name))
            return affected_rows > 0

        except Exception as e:
            self.logger.error(f"移除邮箱标签失败: {e}")
            return False

    def get_emails_by_domain(self, domain: str, limit: int = 100) -> List[EmailModel]:
        """
        根据域名获取邮箱列表

        Args:
            domain: 域名
            limit: 限制数量

        Returns:
            邮箱模型列表
        """
        try:
            query = """
                SELECT * FROM emails
                WHERE domain = ? AND is_active = 1
                ORDER BY created_at DESC
                LIMIT ?
            """
            results = self.db_service.execute_query(query, (domain, limit))

            return [self._row_to_email_model(row) for row in results or []]

        except Exception as e:
            self.logger.error(f"获取域名邮箱列表失败: {e}")
            return []

    def get_emails_by_status(self, status: EmailStatus, limit: int = 100) -> List[EmailModel]:
        """
        根据状态获取邮箱列表

        Args:
            status: 邮箱状态
            limit: 限制数量

        Returns:
            邮箱模型列表
        """
        try:
            query = """
                SELECT * FROM emails
                WHERE status = ? AND is_active = 1
                ORDER BY created_at DESC
                LIMIT ?
            """
            results = self.db_service.execute_query(query, (status.value, limit))

            return [self._row_to_email_model(row) for row in results or []]

        except Exception as e:
            self.logger.error(f"获取状态邮箱列表失败: {e}")
            return []

    def batch_create_emails(self,
                           count: int,
                           prefix_type: str = "random_name",
                           base_prefix: Optional[str] = None,
                           tags: Optional[List[str]] = None,
                           notes: str = "") -> List[EmailModel]:
        """
        批量创建邮箱

        Args:
            count: 创建数量
            prefix_type: 前缀类型
            base_prefix: 基础前缀（用于生成序列）
            tags: 标签列表
            notes: 备注

        Returns:
            创建的邮箱模型列表
        """
        if count <= 0:
            return []

        if count > 100:
            self.logger.warning(f"批量创建数量过大: {count}，限制为100")
            count = 100

        created_emails = []

        try:
            for i in range(count):
                try:
                    # 生成前缀
                    if prefix_type == "sequence" and base_prefix:
                        custom_prefix = f"{base_prefix}_{i+1:03d}"
                        email = self.create_email(
                            prefix_type="custom",
                            custom_prefix=custom_prefix,
                            tags=tags,
                            notes=f"{notes} (批量创建 {i+1}/{count})"
                        )
                    else:
                        email = self.create_email(
                            prefix_type=prefix_type,
                            tags=tags,
                            notes=f"{notes} (批量创建 {i+1}/{count})"
                        )

                    created_emails.append(email)

                    # 添加小延迟确保时间戳唯一性
                    if i < count - 1:
                        import time
                        time.sleep(0.01)

                except Exception as e:
                    self.logger.error(f"批量创建第 {i+1} 个邮箱失败: {e}")
                    continue

            self.logger.info(f"批量创建完成，成功创建 {len(created_emails)} 个邮箱")
            return created_emails

        except Exception as e:
            self.logger.error(f"批量创建邮箱失败: {e}")
            return created_emails

    def batch_delete_emails(self, email_ids: List[int]) -> Dict[str, Any]:
        """
        批量删除邮箱

        Args:
            email_ids: 邮箱ID列表

        Returns:
            删除结果统计
        """
        if not email_ids:
            return {"success": 0, "failed": 0, "total": 0}

        success_count = 0
        failed_count = 0

        try:
            for email_id in email_ids:
                try:
                    if self.delete_email(email_id):
                        success_count += 1
                    else:
                        failed_count += 1
                except Exception as e:
                    self.logger.error(f"删除邮箱 {email_id} 失败: {e}")
                    failed_count += 1

            result = {
                "success": success_count,
                "failed": failed_count,
                "total": len(email_ids)
            }

            self.logger.info(f"批量删除完成: {result}")
            return result

        except Exception as e:
            self.logger.error(f"批量删除邮箱失败: {e}")
            return {"success": success_count, "failed": failed_count, "total": len(email_ids)}

    def get_statistics(self) -> Dict[str, Any]:
        """
        获取邮箱统计信息
        
        Returns:
            统计信息字典
        """
        try:
            stats = {}
            
            # 总邮箱数
            total_query = "SELECT COUNT(*) as count FROM emails WHERE is_active = 1"
            total_result = self.db_service.execute_query(total_query, fetch_one=True)
            stats["total_emails"] = total_result["count"] if total_result else 0
            
            # 按状态统计
            status_query = """
                SELECT status, COUNT(*) as count 
                FROM emails 
                WHERE is_active = 1 
                GROUP BY status
            """
            status_results = self.db_service.execute_query(status_query)
            stats["by_status"] = {row["status"]: row["count"] for row in status_results or []}
            
            # 按域名统计
            domain_query = """
                SELECT domain, COUNT(*) as count 
                FROM emails 
                WHERE is_active = 1 
                GROUP BY domain 
                ORDER BY count DESC 
                LIMIT 10
            """
            domain_results = self.db_service.execute_query(domain_query)
            stats["by_domain"] = {row["domain"]: row["count"] for row in domain_results or []}
            
            # 今日创建数量
            today_query = """
                SELECT COUNT(*) as count 
                FROM emails 
                WHERE is_active = 1 AND DATE(created_at) = DATE('now')
            """
            today_result = self.db_service.execute_query(today_query, fetch_one=True)
            stats["today_created"] = today_result["count"] if today_result else 0
            
            return stats
            
        except Exception as e:
            self.logger.error(f"获取统计信息失败: {e}")
            return {}

    def export_emails(self, format_type: str = "json", filters: Optional[Dict] = None) -> str:
        """
        导出邮箱数据
        
        Args:
            format_type: 导出格式 ("json", "csv")
            filters: 过滤条件
            
        Returns:
            导出的数据字符串
        """
        try:
            # 获取邮箱列表
            emails = self.search_emails(**(filters or {}))
            
            if format_type.lower() == "csv":
                return self._export_to_csv(emails)
            else:
                return self._export_to_json(emails)
                
        except Exception as e:
            self.logger.error(f"导出邮箱数据失败: {e}")
            raise

    def _save_email_to_db(self, email_model: EmailModel) -> int:
        """保存邮箱到数据库"""
        try:
            query = """
                INSERT INTO emails (
                    email_address, domain, prefix, timestamp_suffix,
                    created_at, last_used, updated_at, status,
                    notes, metadata, is_active, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """

            params = (
                email_model.email_address,
                email_model.domain,
                email_model.prefix,
                email_model.timestamp_suffix,
                email_model.created_at.isoformat() if email_model.created_at else None,
                email_model.last_used.isoformat() if email_model.last_used else None,
                email_model.updated_at.isoformat() if email_model.updated_at else None,
                email_model.status.value,
                email_model.notes,
                json.dumps(email_model.metadata) if email_model.metadata else None,
                email_model.is_active,
                email_model.created_by
            )

            with self.db_service.get_cursor() as cursor:
                cursor.execute(query, params)
                email_id = cursor.lastrowid

                # 保存标签关联
                if email_model.tags:
                    self._save_email_tags(cursor, email_id, email_model.tags)

                return email_id

        except Exception as e:
            self.logger.error(f"保存邮箱到数据库失败: {e}")
            raise

    def _update_email_in_db(self, email_model: EmailModel) -> bool:
        """更新数据库中的邮箱信息"""
        try:
            query = """
                UPDATE emails SET
                    last_used = ?,
                    updated_at = ?,
                    status = ?,
                    notes = ?,
                    metadata = ?
                WHERE id = ?
            """

            params = (
                email_model.last_used.isoformat() if email_model.last_used else None,
                email_model.updated_at.isoformat() if email_model.updated_at else None,
                email_model.status.value,
                email_model.notes,
                json.dumps(email_model.metadata) if email_model.metadata else None,
                email_model.id
            )

            affected_rows = self.db_service.execute_update(query, params)
            return affected_rows > 0

        except Exception as e:
            self.logger.error(f"更新邮箱信息失败: {e}")
            return False

    def _save_email_tags(self, cursor, email_id: int, tags: List[str]):
        """保存邮箱标签关联"""
        for tag_name in tags:
            # 确保标签存在
            cursor.execute("SELECT id FROM tags WHERE name = ?", (tag_name,))
            tag_result = cursor.fetchone()

            if tag_result:
                tag_id = tag_result["id"]
            else:
                # 创建新标签
                cursor.execute(
                    "INSERT INTO tags (name, description) VALUES (?, ?)",
                    (tag_name, f"自动创建的标签: {tag_name}")
                )
                tag_id = cursor.lastrowid

            # 创建关联
            cursor.execute(
                "INSERT OR IGNORE INTO email_tags (email_id, tag_id) VALUES (?, ?)",
                (email_id, tag_id)
            )

    def _row_to_email_model(self, row) -> EmailModel:
        """将数据库行转换为邮箱模型"""
        try:
            # 解析时间字段
            def parse_datetime(dt_str):
                if dt_str:
                    try:
                        return datetime.fromisoformat(dt_str)
                    except (ValueError, TypeError):
                        pass
                return None

            # 解析状态
            status = EmailStatus.ACTIVE
            if row["status"]:
                try:
                    status = EmailStatus(row["status"])
                except ValueError:
                    pass

            # 解析元数据
            metadata = {}
            if row["metadata"]:
                try:
                    metadata = json.loads(row["metadata"])
                except (json.JSONDecodeError, TypeError):
                    pass

            # 获取标签
            tags = self._get_email_tags(row["id"])

            return EmailModel(
                id=row["id"],
                email_address=row["email_address"],
                domain=row["domain"],
                prefix=row["prefix"],
                timestamp_suffix=row["timestamp_suffix"] or "",
                created_at=parse_datetime(row["created_at"]),
                last_used=parse_datetime(row["last_used"]),
                updated_at=parse_datetime(row["updated_at"]),
                status=status,
                tags=tags,
                notes=row["notes"] or "",
                metadata=metadata,
                is_active=bool(row["is_active"]),
                created_by=row["created_by"] or "system"
            )

        except Exception as e:
            self.logger.error(f"转换数据库行失败: {e}")
            raise

    def _get_email_tags(self, email_id: int) -> List[str]:
        """获取邮箱的标签列表"""
        try:
            query = """
                SELECT t.name
                FROM tags t
                JOIN email_tags et ON t.id = et.tag_id
                WHERE et.email_id = ?
            """
            results = self.db_service.execute_query(query, (email_id,))
            return [row["name"] for row in results or []]

        except Exception as e:
            self.logger.error(f"获取邮箱标签失败: {e}")
            return []

    def _export_to_json(self, emails: List[EmailModel]) -> str:
        """导出为JSON格式"""
        data = [email.to_dict() for email in emails]
        return json.dumps(data, ensure_ascii=False, indent=2)

    def _export_to_csv(self, emails: List[EmailModel]) -> str:
        """导出为CSV格式"""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # 写入标题行
        headers = [
            "ID", "邮箱地址", "域名", "前缀", "创建时间", "状态",
            "标签", "备注"
        ]
        writer.writerow(headers)

        # 写入数据行
        for email in emails:
            row = [
                email.id,
                email.email_address,
                email.domain,
                email.prefix,
                email.created_at.isoformat() if email.created_at else "",
                email.status_display,
                ",".join(email.tags),
                email.notes
            ]
            writer.writerow(row)

        return output.getvalue()

    # ==================== Phase 3A: 高级搜索和筛选功能 ====================

    def advanced_search_emails(self,
                              keyword: str = "",
                              domain: str = "",
                              status: Optional[EmailStatus] = None,
                              tags: Optional[List[str]] = None,
                              date_from: Optional[str] = None,
                              date_to: Optional[str] = None,
                              created_by: str = "",
                              has_notes: Optional[bool] = None,
                              page: int = 1,
                              page_size: int = 20,
                              sort_by: str = "created_at",
                              sort_order: str = "desc") -> Dict[str, Any]:
        """
        高级搜索邮箱（支持分页和多条件筛选）

        Args:
            keyword: 搜索关键词（邮箱地址、备注）
            domain: 域名筛选
            status: 状态筛选
            tags: 标签筛选（包含任一标签）
            date_from: 开始日期 (YYYY-MM-DD)
            date_to: 结束日期 (YYYY-MM-DD)
            created_by: 创建者筛选
            has_notes: 是否有备注
            page: 页码（从1开始）
            page_size: 每页大小
            sort_by: 排序字段 ("created_at", "email_address", "domain", "status")
            sort_order: 排序方向 ("asc", "desc")

        Returns:
            搜索结果和分页信息
        """
        try:
            # 计算偏移量
            offset = (page - 1) * page_size

            # 构建查询条件
            where_conditions = ["e.is_active = 1"]
            params = []

            if keyword:
                where_conditions.append("(e.email_address LIKE ? OR e.notes LIKE ?)")
                params.extend([f"%{keyword}%", f"%{keyword}%"])

            if domain:
                where_conditions.append("e.domain = ?")
                params.append(domain)

            if status:
                where_conditions.append("e.status = ?")
                params.append(status.value)

            if created_by:
                where_conditions.append("e.created_by = ?")
                params.append(created_by)

            if has_notes is not None:
                if has_notes:
                    where_conditions.append("e.notes IS NOT NULL AND e.notes != ''")
                else:
                    where_conditions.append("(e.notes IS NULL OR e.notes = '')")

            if date_from:
                where_conditions.append("DATE(e.created_at) >= ?")
                params.append(date_from)

            if date_to:
                where_conditions.append("DATE(e.created_at) <= ?")
                params.append(date_to)

            # 标签筛选
            if tags:
                tag_conditions = []
                for tag in tags:
                    tag_conditions.append("""
                        EXISTS (
                            SELECT 1 FROM email_tags et
                            JOIN tags t ON et.tag_id = t.id
                            WHERE et.email_id = e.id AND t.name = ?
                        )
                    """)
                    params.append(tag)
                where_conditions.append(f"({' OR '.join(tag_conditions)})")

            where_clause = " AND ".join(where_conditions)

            # 构建排序子句
            sort_column = "e.created_at"
            if sort_by == "email_address":
                sort_column = "e.email_address"
            elif sort_by == "domain":
                sort_column = "e.domain"
            elif sort_by == "status":
                sort_column = "e.status"

            sort_direction = "ASC" if sort_order.lower() == "asc" else "DESC"

            # 查询总数
            count_query = f"""
                SELECT COUNT(*) as total
                FROM emails e
                WHERE {where_clause}
            """
            count_result = self.db_service.execute_query(count_query, params, fetch_one=True)
            total = count_result["total"] if count_result else 0

            # 查询数据
            data_query = f"""
                SELECT e.*
                FROM emails e
                WHERE {where_clause}
                ORDER BY {sort_column} {sort_direction}
                LIMIT ? OFFSET ?
            """
            params.extend([page_size, offset])

            results = self.db_service.execute_query(data_query, params)

            # 转换为邮箱模型
            emails = [self._row_to_email_model(row) for row in results or []]

            # 计算分页信息
            total_pages = (total + page_size - 1) // page_size

            return {
                "emails": emails,
                "pagination": {
                    "current_page": page,
                    "page_size": page_size,
                    "total_items": total,
                    "total_pages": total_pages,
                    "has_next": page < total_pages,
                    "has_prev": page > 1
                },
                "filters": {
                    "keyword": keyword,
                    "domain": domain,
                    "status": status.value if status else None,
                    "tags": tags,
                    "date_from": date_from,
                    "date_to": date_to,
                    "created_by": created_by,
                    "has_notes": has_notes,
                    "sort_by": sort_by,
                    "sort_order": sort_order
                }
            }

        except Exception as e:
            self.logger.error(f"高级搜索邮箱失败: {e}")
            return {
                "emails": [],
                "pagination": {
                    "current_page": 1,
                    "page_size": page_size,
                    "total_items": 0,
                    "total_pages": 0,
                    "has_next": False,
                    "has_prev": False
                },
                "filters": {}
            }

    def get_emails_by_multiple_tags(self, tag_names: List[str],
                                   match_all: bool = True,
                                   limit: int = 100) -> List[EmailModel]:
        """
        根据多个标签获取邮箱

        Args:
            tag_names: 标签名称列表
            match_all: True=必须包含所有标签，False=包含任一标签
            limit: 限制数量

        Returns:
            邮箱模型列表
        """
        try:
            if not tag_names:
                return []

            if match_all:
                # 必须包含所有标签
                query = """
                    SELECT e.* FROM emails e
                    WHERE e.is_active = 1
                """

                for i, tag_name in enumerate(tag_names):
                    query += f"""
                        AND EXISTS (
                            SELECT 1 FROM email_tags et{i}
                            JOIN tags t{i} ON et{i}.tag_id = t{i}.id
                            WHERE et{i}.email_id = e.id AND t{i}.name = ?
                        )
                    """

                query += " ORDER BY e.created_at DESC LIMIT ?"
                params = tag_names + [limit]
            else:
                # 包含任一标签
                placeholders = ",".join(["?" for _ in tag_names])
                query = f"""
                    SELECT DISTINCT e.* FROM emails e
                    JOIN email_tags et ON e.id = et.email_id
                    JOIN tags t ON et.tag_id = t.id
                    WHERE e.is_active = 1 AND t.name IN ({placeholders})
                    ORDER BY e.created_at DESC
                    LIMIT ?
                """
                params = tag_names + [limit]

            results = self.db_service.execute_query(query, params)
            return [self._row_to_email_model(row) for row in results or []]

        except Exception as e:
            self.logger.error(f"根据多个标签获取邮箱失败: {e}")
            return []

    def get_emails_by_date_range(self,
                                start_date: str,
                                end_date: str,
                                date_field: str = "created_at",
                                limit: int = 100) -> List[EmailModel]:
        """
        根据日期范围获取邮箱

        Args:
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            date_field: 日期字段 ("created_at", "last_used", "updated_at")
            limit: 限制数量

        Returns:
            邮箱模型列表
        """
        try:
            # 验证日期字段
            valid_fields = ["created_at", "last_used", "updated_at"]
            if date_field not in valid_fields:
                date_field = "created_at"

            query = f"""
                SELECT * FROM emails
                WHERE is_active = 1
                AND DATE({date_field}) >= ?
                AND DATE({date_field}) <= ?
                ORDER BY {date_field} DESC
                LIMIT ?
            """

            results = self.db_service.execute_query(query, (start_date, end_date, limit))
            return [self._row_to_email_model(row) for row in results or []]

        except Exception as e:
            self.logger.error(f"根据日期范围获取邮箱失败: {e}")
            return []

    def get_email_statistics_by_period(self,
                                      period: str = "month",
                                      limit: int = 12) -> List[Dict[str, Any]]:
        """
        获取按时间段的邮箱统计

        Args:
            period: 时间段 ("day", "week", "month", "year")
            limit: 限制数量

        Returns:
            统计数据列表
        """
        try:
            # 根据时间段构建查询
            if period == "day":
                date_format = "%Y-%m-%d"
                group_by = "DATE(created_at)"
            elif period == "week":
                date_format = "%Y-W%W"
                group_by = "strftime('%Y-W%W', created_at)"
            elif period == "year":
                date_format = "%Y"
                group_by = "strftime('%Y', created_at)"
            else:  # month
                date_format = "%Y-%m"
                group_by = "strftime('%Y-%m', created_at)"

            query = f"""
                SELECT
                    {group_by} as period,
                    COUNT(*) as total_count,
                    SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_count,
                    SUM(CASE WHEN status = 'inactive' THEN 1 ELSE 0 END) as inactive_count,
                    SUM(CASE WHEN status = 'archived' THEN 1 ELSE 0 END) as archived_count
                FROM emails
                WHERE is_active = 1
                GROUP BY {group_by}
                ORDER BY period DESC
                LIMIT ?
            """

            results = self.db_service.execute_query(query, (limit,))

            return [
                {
                    "period": row["period"],
                    "total_count": row["total_count"],
                    "active_count": row["active_count"],
                    "inactive_count": row["inactive_count"],
                    "archived_count": row["archived_count"]
                }
                for row in results or []
            ]

        except Exception as e:
            self.logger.error(f"获取时间段统计失败: {e}")
            return []

    def search_emails_with_filters(self, filters: Dict[str, Any]) -> List[EmailModel]:
        """
        使用过滤器字典搜索邮箱

        Args:
            filters: 过滤条件字典

        Returns:
            邮箱模型列表
        """
        try:
            # 提取过滤条件
            keyword = filters.get("keyword", "")
            domain = filters.get("domain", "")
            status = filters.get("status")
            tags = filters.get("tags", [])
            date_from = filters.get("date_from")
            date_to = filters.get("date_to")
            created_by = filters.get("created_by", "")
            has_notes = filters.get("has_notes")
            limit = filters.get("limit", 100)

            # 转换状态
            email_status = None
            if status:
                try:
                    email_status = EmailStatus(status)
                except ValueError:
                    pass

            # 使用高级搜索
            result = self.advanced_search_emails(
                keyword=keyword,
                domain=domain,
                status=email_status,
                tags=tags,
                date_from=date_from,
                date_to=date_to,
                created_by=created_by,
                has_notes=has_notes,
                page=1,
                page_size=limit
            )

            return result.get("emails", [])

        except Exception as e:
            self.logger.error(f"使用过滤器搜索邮箱失败: {e}")
            return []

    # ==================== Phase 3A: 高级数据导出功能 ====================

    def export_emails_advanced(self,
                              format_type: str = "json",
                              filters: Optional[Dict[str, Any]] = None,
                              fields: Optional[List[str]] = None,
                              include_tags: bool = True,
                              include_metadata: bool = False) -> str:
        """
        高级邮箱数据导出

        Args:
            format_type: 导出格式 ("json", "csv", "xlsx")
            filters: 过滤条件
            fields: 要导出的字段列表
            include_tags: 是否包含标签信息
            include_metadata: 是否包含元数据

        Returns:
            导出的数据字符串
        """
        try:
            # 获取邮箱列表
            emails = self.search_emails_with_filters(filters or {})

            if format_type.lower() == "csv":
                return self._export_to_csv_advanced(emails, fields, include_tags, include_metadata)
            elif format_type.lower() == "xlsx":
                return self._export_to_xlsx(emails, fields, include_tags, include_metadata)
            else:
                return self._export_to_json_advanced(emails, fields, include_tags, include_metadata)

        except Exception as e:
            self.logger.error(f"高级导出邮箱数据失败: {e}")
            raise

    def _export_to_json_advanced(self,
                                emails: List[EmailModel],
                                fields: Optional[List[str]] = None,
                                include_tags: bool = True,
                                include_metadata: bool = False) -> str:
        """高级JSON导出"""
        import json

        try:
            # 默认字段
            default_fields = [
                "id", "email_address", "domain", "prefix", "timestamp_suffix",
                "created_at", "last_used", "updated_at", "status", "notes", "created_by"
            ]

            export_fields = fields or default_fields
            export_data = []

            for email in emails:
                email_data = {}

                # 导出指定字段
                for field in export_fields:
                    if hasattr(email, field):
                        value = getattr(email, field)
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        elif hasattr(value, 'value'):  # Enum类型
                            value = value.value
                        email_data[field] = value

                # 包含标签
                if include_tags:
                    email_data["tags"] = email.tags

                # 包含元数据
                if include_metadata and email.metadata:
                    email_data["metadata"] = email.metadata

                export_data.append(email_data)

            return json.dumps(export_data, ensure_ascii=False, indent=2)

        except Exception as e:
            self.logger.error(f"JSON高级导出失败: {e}")
            return ""

    def _export_to_csv_advanced(self,
                               emails: List[EmailModel],
                               fields: Optional[List[str]] = None,
                               include_tags: bool = True,
                               include_metadata: bool = False) -> str:
        """高级CSV导出"""
        import csv
        import io

        try:
            output = io.StringIO()

            # 默认字段
            default_fields = [
                "id", "email_address", "domain", "prefix", "timestamp_suffix",
                "created_at", "last_used", "updated_at", "status", "notes", "created_by"
            ]

            export_fields = fields or default_fields

            # 添加标签和元数据字段
            if include_tags:
                export_fields.append("tags")
            if include_metadata:
                export_fields.append("metadata")

            writer = csv.DictWriter(output, fieldnames=export_fields)
            writer.writeheader()

            for email in emails:
                row_data = {}

                # 导出指定字段
                for field in export_fields:
                    if field == "tags":
                        row_data[field] = ",".join(email.tags) if email.tags else ""
                    elif field == "metadata":
                        row_data[field] = json.dumps(email.metadata) if email.metadata else ""
                    elif hasattr(email, field):
                        value = getattr(email, field)
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        elif hasattr(value, 'value'):  # Enum类型
                            value = value.value
                        row_data[field] = value
                    else:
                        row_data[field] = ""

                writer.writerow(row_data)

            return output.getvalue()

        except Exception as e:
            self.logger.error(f"CSV高级导出失败: {e}")
            return ""

    def _export_to_xlsx(self,
                       emails: List[EmailModel],
                       fields: Optional[List[str]] = None,
                       include_tags: bool = True,
                       include_metadata: bool = False) -> bytes:
        """导出为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "邮箱数据"

            # 默认字段
            default_fields = [
                "id", "email_address", "domain", "prefix", "timestamp_suffix",
                "created_at", "last_used", "updated_at", "status", "notes", "created_by"
            ]

            export_fields = fields or default_fields

            # 添加标签和元数据字段
            if include_tags:
                export_fields.append("tags")
            if include_metadata:
                export_fields.append("metadata")

            # 写入表头
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            for col, field in enumerate(export_fields, 1):
                cell = ws.cell(row=1, column=col, value=field)
                cell.font = header_font
                cell.fill = header_fill

            # 写入数据
            for row, email in enumerate(emails, 2):
                for col, field in enumerate(export_fields, 1):
                    if field == "tags":
                        value = ",".join(email.tags) if email.tags else ""
                    elif field == "metadata":
                        value = json.dumps(email.metadata) if email.metadata else ""
                    elif hasattr(email, field):
                        value = getattr(email, field)
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        elif hasattr(value, 'value'):  # Enum类型
                            value = value.value
                    else:
                        value = ""

                    ws.cell(row=row, column=col, value=value)

            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 保存到字节流
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return output.getvalue()

        except ImportError:
            self.logger.error("openpyxl库未安装，无法导出Excel格式")
            raise ValueError("Excel导出功能需要安装openpyxl库")
        except Exception as e:
            self.logger.error(f"Excel导出失败: {e}")
            raise
