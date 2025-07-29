# -*- coding: utf-8 -*-
"""
域名邮箱管理器 - 数据导出服务
专门处理各种格式的数据导出功能
"""

import json
import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from pathlib import Path

from models.email_model import EmailModel
from models.tag_model import TagModel
from services.database_service import DatabaseService
from services.email_service import EmailService
from services.tag_service import TagService
from utils.logger import get_logger


class ExportService:
    """
    数据导出服务类
    
    提供邮箱、标签等数据的多格式导出功能
    """

    def __init__(self, db_service: DatabaseService):
        """
        初始化导出服务
        
        Args:
            db_service: 数据库服务实例
        """
        self.db_service = db_service
        self.logger = get_logger(__name__)
        
        # 初始化其他服务
        self.email_service = None
        self.tag_service = None
        
        self.logger.info("数据导出服务初始化完成")

    def set_services(self, email_service: EmailService, tag_service: TagService):
        """
        设置依赖的服务实例
        
        Args:
            email_service: 邮箱服务实例
            tag_service: 标签服务实例
        """
        self.email_service = email_service
        self.tag_service = tag_service

    def export_all_data(self, 
                       format_type: str = "json",
                       output_path: Optional[str] = None,
                       include_deleted: bool = False) -> Union[str, bytes]:
        """
        导出所有数据（邮箱、标签、配置等）
        
        Args:
            format_type: 导出格式 ("json", "csv", "xlsx")
            output_path: 输出文件路径（可选）
            include_deleted: 是否包含已删除的数据
            
        Returns:
            导出的数据内容
        """
        try:
            # 收集所有数据
            export_data = {
                "export_info": {
                    "timestamp": datetime.now().isoformat(),
                    "format": format_type,
                    "include_deleted": include_deleted
                },
                "emails": self._get_emails_for_export(include_deleted),
                "tags": self._get_tags_for_export(include_deleted),
                "statistics": self._get_export_statistics()
            }
            
            # 根据格式导出
            if format_type.lower() == "json":
                result = self._export_all_to_json(export_data)
            elif format_type.lower() == "csv":
                result = self._export_all_to_csv(export_data)
            elif format_type.lower() == "xlsx":
                result = self._export_all_to_xlsx(export_data)
            else:
                raise ValueError(f"不支持的导出格式: {format_type}")
            
            # 保存到文件（如果指定了路径）
            if output_path:
                self._save_to_file(result, output_path, format_type)
            
            self.logger.info(f"成功导出所有数据，格式: {format_type}")
            return result
            
        except Exception as e:
            self.logger.error(f"导出所有数据失败: {e}")
            raise

    def export_emails_with_template(self, 
                                   template_name: str,
                                   filters: Optional[Dict[str, Any]] = None) -> str:
        """
        使用预定义模板导出邮箱数据
        
        Args:
            template_name: 模板名称 ("simple", "detailed", "report")
            filters: 过滤条件
            
        Returns:
            导出的数据字符串
        """
        try:
            # 获取邮箱数据
            if self.email_service:
                emails = self.email_service.search_emails_with_filters(filters or {})
            else:
                emails = []
            
            # 根据模板生成数据
            if template_name == "simple":
                return self._export_simple_template(emails)
            elif template_name == "detailed":
                return self._export_detailed_template(emails)
            elif template_name == "report":
                return self._export_report_template(emails)
            else:
                raise ValueError(f"未知的模板: {template_name}")
                
        except Exception as e:
            self.logger.error(f"使用模板导出失败: {e}")
            raise

    def _get_emails_for_export(self, include_deleted: bool = False) -> List[Dict[str, Any]]:
        """获取用于导出的邮箱数据"""
        try:
            # 构建查询
            if include_deleted:
                query = "SELECT * FROM emails ORDER BY created_at DESC"
                params = ()
            else:
                query = "SELECT * FROM emails WHERE is_active = 1 ORDER BY created_at DESC"
                params = ()
            
            results = self.db_service.execute_query(query, params)
            
            emails_data = []
            for row in results or []:
                # 获取标签
                tag_query = """
                    SELECT t.name FROM tags t
                    JOIN email_tags et ON t.id = et.tag_id
                    WHERE et.email_id = ?
                """
                tag_results = self.db_service.execute_query(tag_query, (row["id"],))
                tags = [tag_row["name"] for tag_row in tag_results or []]
                
                email_data = {
                    "id": row["id"],
                    "email_address": row["email_address"],
                    "domain": row["domain"],
                    "prefix": row["prefix"],
                    "timestamp_suffix": row["timestamp_suffix"],
                    "created_at": row["created_at"],
                    "last_used": row["last_used"],
                    "updated_at": row["updated_at"],
                    "status": row["status"],
                    "notes": row["notes"],
                    "is_active": row["is_active"],
                    "created_by": row["created_by"],
                    "tags": tags,
                    "metadata": json.loads(row["metadata"]) if row["metadata"] else {}
                }
                emails_data.append(email_data)
            
            return emails_data
            
        except Exception as e:
            self.logger.error(f"获取邮箱导出数据失败: {e}")
            return []

    def _get_tags_for_export(self, include_deleted: bool = False) -> List[Dict[str, Any]]:
        """获取用于导出的标签数据"""
        try:
            # 构建查询
            if include_deleted:
                query = "SELECT * FROM tags ORDER BY created_at DESC"
                params = ()
            else:
                query = "SELECT * FROM tags WHERE is_active = 1 ORDER BY created_at DESC"
                params = ()
            
            results = self.db_service.execute_query(query, params)
            
            tags_data = []
            for row in results or []:
                # 获取使用统计
                usage_query = """
                    SELECT COUNT(*) as count FROM email_tags et
                    JOIN emails e ON et.email_id = e.id
                    WHERE et.tag_id = ? AND e.is_active = 1
                """
                usage_result = self.db_service.execute_query(usage_query, (row["id"],), fetch_one=True)
                usage_count = usage_result["count"] if usage_result else 0
                
                tag_data = {
                    "id": row["id"],
                    "name": row["name"],
                    "description": row["description"],
                    "color": row["color"],
                    "icon": row["icon"],
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                    "is_system": row["is_system"],
                    "is_active": row["is_active"],
                    "usage_count": usage_count
                }
                tags_data.append(tag_data)
            
            return tags_data
            
        except Exception as e:
            self.logger.error(f"获取标签导出数据失败: {e}")
            return []

    def _get_export_statistics(self) -> Dict[str, Any]:
        """获取导出统计信息"""
        try:
            stats = {}
            
            # 邮箱统计
            email_stats = self.db_service.execute_query(
                "SELECT COUNT(*) as total, SUM(CASE WHEN is_active = 1 THEN 1 ELSE 0 END) as active FROM emails",
                fetch_one=True
            )
            stats["emails"] = {
                "total": email_stats["total"] if email_stats else 0,
                "active": email_stats["active"] if email_stats else 0
            }
            
            # 标签统计
            tag_stats = self.db_service.execute_query(
                "SELECT COUNT(*) as total, SUM(CASE WHEN is_active = 1 THEN 1 ELSE 0 END) as active FROM tags",
                fetch_one=True
            )
            stats["tags"] = {
                "total": tag_stats["total"] if tag_stats else 0,
                "active": tag_stats["active"] if tag_stats else 0
            }
            
            # 关联统计
            relation_stats = self.db_service.execute_query(
                "SELECT COUNT(*) as total FROM email_tags",
                fetch_one=True
            )
            stats["email_tag_relations"] = relation_stats["total"] if relation_stats else 0
            
            return stats
            
        except Exception as e:
            self.logger.error(f"获取导出统计失败: {e}")
            return {}

    def _export_all_to_json(self, data: Dict[str, Any]) -> str:
        """导出所有数据为JSON格式"""
        try:
            return json.dumps(data, ensure_ascii=False, indent=2)
        except Exception as e:
            self.logger.error(f"JSON导出失败: {e}")
            return ""

    def _export_simple_template(self, emails: List[EmailModel]) -> str:
        """简单模板导出"""
        try:
            simple_data = []
            for email in emails:
                simple_data.append({
                    "email_address": email.email_address,
                    "domain": email.domain,
                    "status": email.status.value,
                    "created_at": email.created_at.isoformat() if email.created_at else None,
                    "tags": email.tags
                })
            
            return json.dumps(simple_data, ensure_ascii=False, indent=2)
            
        except Exception as e:
            self.logger.error(f"简单模板导出失败: {e}")
            return ""

    def _export_detailed_template(self, emails: List[EmailModel]) -> str:
        """详细模板导出"""
        try:
            detailed_data = {
                "export_info": {
                    "timestamp": datetime.now().isoformat(),
                    "total_emails": len(emails),
                    "template": "detailed"
                },
                "emails": []
            }

            for email in emails:
                email_data = {
                    "id": email.id,
                    "email_address": email.email_address,
                    "domain": email.domain,
                    "prefix": email.prefix,
                    "timestamp_suffix": email.timestamp_suffix,
                    "created_at": email.created_at.isoformat() if email.created_at else None,
                    "last_used": email.last_used.isoformat() if email.last_used else None,
                    "updated_at": email.updated_at.isoformat() if email.updated_at else None,
                    "status": email.status.value,
                    "notes": email.notes,
                    "tags": email.tags,
                    "metadata": email.metadata,
                    "created_by": email.created_by
                }
                detailed_data["emails"].append(email_data)

            return json.dumps(detailed_data, ensure_ascii=False, indent=2)

        except Exception as e:
            self.logger.error(f"详细模板导出失败: {e}")
            return ""

    def _export_report_template(self, emails: List[EmailModel]) -> str:
        """报告模板导出"""
        try:
            # 统计分析
            total_emails = len(emails)
            status_stats = {}
            domain_stats = {}
            tag_stats = {}

            for email in emails:
                # 状态统计
                status = email.status.value
                status_stats[status] = status_stats.get(status, 0) + 1

                # 域名统计
                domain = email.domain
                domain_stats[domain] = domain_stats.get(domain, 0) + 1

                # 标签统计
                for tag in email.tags:
                    tag_stats[tag] = tag_stats.get(tag, 0) + 1

            report_data = {
                "report_info": {
                    "generated_at": datetime.now().isoformat(),
                    "template": "report",
                    "total_emails": total_emails
                },
                "statistics": {
                    "by_status": status_stats,
                    "by_domain": domain_stats,
                    "by_tags": dict(sorted(tag_stats.items(), key=lambda x: x[1], reverse=True)[:10])
                },
                "summary": {
                    "most_used_domain": max(domain_stats.items(), key=lambda x: x[1])[0] if domain_stats else None,
                    "most_used_tag": max(tag_stats.items(), key=lambda x: x[1])[0] if tag_stats else None,
                    "active_emails": status_stats.get("active", 0),
                    "inactive_emails": status_stats.get("inactive", 0),
                    "archived_emails": status_stats.get("archived", 0)
                },
                "emails": [
                    {
                        "email_address": email.email_address,
                        "domain": email.domain,
                        "status": email.status.value,
                        "created_at": email.created_at.isoformat() if email.created_at else None,
                        "tags_count": len(email.tags),
                        "has_notes": bool(email.notes)
                    }
                    for email in emails
                ]
            }

            return json.dumps(report_data, ensure_ascii=False, indent=2)

        except Exception as e:
            self.logger.error(f"报告模板导出失败: {e}")
            return ""

    def _export_all_to_csv(self, data: Dict[str, Any]) -> str:
        """导出所有数据为CSV格式（多个表）"""
        try:
            output = io.StringIO()

            # 导出信息
            output.write("# 数据导出信息\n")
            export_info = data.get("export_info", {})
            for key, value in export_info.items():
                output.write(f"# {key}: {value}\n")
            output.write("\n")

            # 邮箱数据
            output.write("# 邮箱数据\n")
            emails = data.get("emails", [])
            if emails:
                email_fields = ["id", "email_address", "domain", "prefix", "created_at", "status", "tags", "notes"]
                writer = csv.DictWriter(output, fieldnames=email_fields)
                writer.writeheader()

                for email in emails:
                    row = {field: email.get(field, "") for field in email_fields}
                    if isinstance(row["tags"], list):
                        row["tags"] = ",".join(row["tags"])
                    writer.writerow(row)

            output.write("\n")

            # 标签数据
            output.write("# 标签数据\n")
            tags = data.get("tags", [])
            if tags:
                tag_fields = ["id", "name", "description", "color", "icon", "created_at", "usage_count"]
                writer = csv.DictWriter(output, fieldnames=tag_fields)
                writer.writeheader()

                for tag in tags:
                    row = {field: tag.get(field, "") for field in tag_fields}
                    writer.writerow(row)

            return output.getvalue()

        except Exception as e:
            self.logger.error(f"CSV导出失败: {e}")
            return ""

    def _export_all_to_xlsx(self, data: Dict[str, Any]) -> bytes:
        """导出所有数据为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()

            # 删除默认工作表
            wb.remove(wb.active)

            # 创建邮箱数据工作表
            emails_ws = wb.create_sheet("邮箱数据")
            emails = data.get("emails", [])

            if emails:
                # 邮箱表头
                email_headers = ["ID", "邮箱地址", "域名", "前缀", "创建时间", "状态", "标签", "备注"]
                for col, header in enumerate(email_headers, 1):
                    cell = emails_ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # 邮箱数据
                for row, email in enumerate(emails, 2):
                    emails_ws.cell(row=row, column=1, value=email.get("id", ""))
                    emails_ws.cell(row=row, column=2, value=email.get("email_address", ""))
                    emails_ws.cell(row=row, column=3, value=email.get("domain", ""))
                    emails_ws.cell(row=row, column=4, value=email.get("prefix", ""))
                    emails_ws.cell(row=row, column=5, value=email.get("created_at", ""))
                    emails_ws.cell(row=row, column=6, value=email.get("status", ""))
                    emails_ws.cell(row=row, column=7, value=",".join(email.get("tags", [])))
                    emails_ws.cell(row=row, column=8, value=email.get("notes", ""))

            # 创建标签数据工作表
            tags_ws = wb.create_sheet("标签数据")
            tags = data.get("tags", [])

            if tags:
                # 标签表头
                tag_headers = ["ID", "名称", "描述", "颜色", "图标", "创建时间", "使用次数"]
                for col, header in enumerate(tag_headers, 1):
                    cell = tags_ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # 标签数据
                for row, tag in enumerate(tags, 2):
                    tags_ws.cell(row=row, column=1, value=tag.get("id", ""))
                    tags_ws.cell(row=row, column=2, value=tag.get("name", ""))
                    tags_ws.cell(row=row, column=3, value=tag.get("description", ""))
                    tags_ws.cell(row=row, column=4, value=tag.get("color", ""))
                    tags_ws.cell(row=row, column=5, value=tag.get("icon", ""))
                    tags_ws.cell(row=row, column=6, value=tag.get("created_at", ""))
                    tags_ws.cell(row=row, column=7, value=tag.get("usage_count", 0))

            # 保存到字节流
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return output.getvalue()

        except ImportError:
            self.logger.error("openpyxl库未安装，无法导出Excel格式")
            raise ValueError("Excel导出功能需要安装openpyxl库")
        except Exception as e:
            self.logger.error(f"Excel导出失败: {e}")
            raise

    def _save_to_file(self, data: Union[str, bytes], file_path: str, format_type: str):
        """保存数据到文件"""
        try:
            path = Path(file_path)
            path.parent.mkdir(parents=True, exist_ok=True)

            if isinstance(data, bytes):
                with open(path, 'wb') as f:
                    f.write(data)
            else:
                with open(path, 'w', encoding='utf-8') as f:
                    f.write(data)

            self.logger.info(f"数据已保存到文件: {file_path}")

        except Exception as e:
            self.logger.error(f"保存文件失败: {e}")
            raise
